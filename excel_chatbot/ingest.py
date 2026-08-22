"""
Step 1: Ingest Excel sheets -> Parquet -> DuckDB.

Reads each sheet of a large .xlsx file in chunks (to avoid loading
1M rows x 4000 cols fully into memory via pandas at once), writes it
out as Parquet, then registers it as a table in a local DuckDB file.

Usage:
    python ingest.py --xlsx path/to/workbook.xlsx --outdir data --db data/warehouse.duckdb
"""

import argparse
import os
import duckdb
import pandas as pd
from openpyxl import load_workbook


def sanitize_col(name: str, idx: int) -> str:
    """Make a safe, unique-ish SQL column name from an Excel header."""
    name = str(name).strip() if name is not None else f"col_{idx}"
    name = "".join(c if c.isalnum() or c == "_" else "_" for c in name)
    if not name or name[0].isdigit():
        name = f"c_{name}"
    return name.lower()


def sheet_to_parquet(xlsx_path: str, sheet_name: str, out_path: str, chunk_rows: int = 50_000):
    """
    Stream a single sheet out to Parquet in chunks using openpyxl's
    read-only mode, so we never hold the full 1M x 4000 sheet in RAM.
    """
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb[sheet_name]

    rows_iter = ws.iter_rows(values_only=True)
    header = next(rows_iter)
    columns = [sanitize_col(h, i) for i, h in enumerate(header)]

    writer = None
    buffer = []

    import pyarrow as pa
    import pyarrow.parquet as pq

    def flush(buffer):
        nonlocal writer
        if not buffer:
            return
        df = pd.DataFrame(buffer, columns=columns)
        table = pa.Table.from_pandas(df, preserve_index=False)
        if writer is None:
            writer = pq.ParquetWriter(out_path, table.schema)
        writer.write_table(table)
        buffer.clear()

    for row in rows_iter:
        buffer.append(row)
        if len(buffer) >= chunk_rows:
            flush(buffer)

    flush(buffer)
    if writer is not None:
        writer.close()

    wb.close()
    return columns


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--xlsx", required=True, help="Path to the source .xlsx workbook")
    ap.add_argument("--outdir", default="data", help="Directory to write Parquet files")
    ap.add_argument("--db", default="data/warehouse.duckdb", help="Path to DuckDB file")
    ap.add_argument("--chunk-rows", type=int, default=50_000)
    args = ap.parse_args()

    os.makedirs(args.outdir, exist_ok=True)

    wb = load_workbook(args.xlsx, read_only=True)
    sheet_names = wb.sheetnames
    wb.close()

    con = duckdb.connect(args.db)

    for sheet in sheet_names:
        table_name = sanitize_col(sheet, 0)
        parquet_path = os.path.join(args.outdir, f"{table_name}.parquet")
        print(f"Converting sheet '{sheet}' -> {parquet_path} ...")
        sheet_to_parquet(args.xlsx, sheet, parquet_path, chunk_rows=args.chunk_rows)

        print(f"Registering table '{table_name}' in DuckDB ...")
        con.execute(f"""
            CREATE OR REPLACE TABLE {table_name} AS
            SELECT * FROM read_parquet('{parquet_path}')
        """)

        n_rows = con.execute(f"SELECT COUNT(*) FROM {table_name}").fetchone()[0]
        n_cols = len(con.execute(f"DESCRIBE {table_name}").fetchall())
        print(f"  -> {n_rows:,} rows, {n_cols:,} columns")

    con.close()
    print(f"\nDone. DuckDB file: {args.db}")


if __name__ == "__main__":
    main()
